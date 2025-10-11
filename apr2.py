import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers


def calculate_apr_from_monthly(P, A, n, tol=1e-10, max_iter=10000):
    """
    Calculate APR using binary search, with improved stability.
    Returns (monthly_rate, annual_percentage_rate)
    """
    if A * n <= P:
        print("\n⚠️ Invalid loan terms: monthly payments too low to repay principal.")
        return 0.0, 0.0

    low, high = 0.0, 2.0  # up to 200% monthly
    r = 0.0

    for _ in range(max_iter):
        r = (low + high) / 2
        denom = ((1 + r) ** n - 1)
        if denom == 0:
            denom = 1e-12
        estimated_payment = P * r * (1 + r) ** n / denom

        diff = estimated_payment - A

        if abs(diff) < tol:
            break
        if diff > 0:
            high = r
        else:
            low = r

    apr = r * 12 * 100
    return r, apr


def generate_amortization_schedule(P, A, r, n):
    """Generate amortization schedule."""
    schedule = []
    balance = P

    for month in range(1, n + 1):
        interest = balance * r
        principal = A - interest
        balance -= principal
        balance = max(balance, 0)
        schedule.append({
            "Month": month,
            "Payment": round(A, 2),
            "Interest": round(interest, 2),
            "Principal": round(principal, 2),
            "Remaining Balance": round(balance, 2)
        })
    return schedule


def display_schedule(schedule):
    """Print amortization schedule to console."""
    print("\nAmortization Schedule")
    print("-" * 70)
    print(f"{'Month':<6}{'Payment':>12}{'Interest':>12}{'Principal':>12}{'Balance':>15}")
    print("-" * 70)
    for row in schedule:
        print(f"{row['Month']:<6}{row['Payment']:>12.2f}{row['Interest']:>12.2f}"
              f"{row['Principal']:>12.2f}{row['Remaining Balance']:>15.2f}")


def export_schedule(schedule, P, A, n, apr, output_folder="."):
    """Export schedule and summary to Excel and CSV, with optional output folder."""
    df = pd.DataFrame(schedule)

    # --- Totals ---
    total_payment = df["Payment"].sum()
    total_interest = df["Interest"].sum()
    total_principal = df["Principal"].sum()

    totals_row = {
        "Month": "TOTAL",
        "Payment": round(total_payment, 2),
        "Interest": round(total_interest, 2),
        "Principal": round(total_principal, 2),
        "Remaining Balance": ""
    }
    df.loc[len(df)] = totals_row

    # --- Loan summary with totals ---
    total_cost = total_payment
    summary_data = {
        "Loan Detail": [
            "Loan Amount (Principal)",
            "Monthly Payment",
            "Number of Monthly Payments",
            "Annual Percentage Rate (APR)",
            "Total Interest Paid",
            "Total Cost of Loan (Principal + Interest)"
        ],
        "Value": [P, A, n, f"{apr:.2f}%", total_interest, total_cost]
    }

    # --- Prepare filenames and ensure output folder exists ---
    apr_str = f"{apr:.2f}".replace(".", "_")
    filename_base = f"Loan_{int(P)}_{n}mo_{apr_str}APR"
    os.makedirs(output_folder, exist_ok=True)

    csv_file = os.path.join(output_folder, f"{filename_base}.csv")
    excel_file = os.path.join(output_folder, f"{filename_base}.xlsx")

    # --- Save CSV ---
    df.to_csv(csv_file, index=False)

    # --- Save Excel ---
    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Amortization Schedule", index=False, startrow=0)
        df.to_excel(writer, sheet_name="Amortization Schedule", index=False, startrow=9)

    # --- Format Excel ---
    wb = load_workbook(excel_file)
    ws = wb["Amortization Schedule"]

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    border = Border(bottom=Side(style="thin"))
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Format summary section
    for cell in ws["A"][0:6]:
        cell.font = bold_font
    for cell in ws["B"][0:6]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    ws["B4"].alignment = center_align  # APR

    # Header formatting
    header_row = 10
    for cell in ws[header_row]:
        cell.font = bold_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    # Format numeric columns
    for row in ws.iter_rows(min_row=header_row + 1, min_col=2, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Highlight totals row
    total_row = ws.max_row
    for cell in ws[total_row]:
        cell.font = Font(bold=True, color="1F497D")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Auto column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col[0].column_letter].width = max_length + 3

    wb.save(excel_file)

    print(f"\n✅ Schedule exported successfully:")
    print(f"  - {excel_file}")
    print(f"  - {csv_file}")


# ---------------- MAIN PROGRAM ----------------
if __name__ == "__main__":
    print("📘 Loan APR Calculator & Amortization Schedule")
    print("------------------------------------------------")

    # --- Command-line mode ---
    if len(sys.argv) >= 4:
        try:
            P = float(sys.argv[1])
            A = float(sys.argv[2])
            n = int(sys.argv[3])
            output_folder = sys.argv[4] if len(sys.argv) >= 5 else "."
            print(f"Using command-line inputs: Loan={P}, Payment={A}, Months={n}")
            if output_folder != ".":
                print(f"Output folder: {output_folder}")
        except ValueError:
            print("❌ Invalid command-line arguments. Please provide numeric values.")
            sys.exit(1)
    else:
        # --- Interactive mode ---
        P = float(input("Enter the loan amount (principal): "))
        A = float(input("Enter the monthly payment: "))
        n = int(input("Enter the number of monthly payments: "))
        output_folder = input("Enter output folder path (or leave blank for current folder): ").strip() or "."

    # --- Calculation ---
    monthly_rate, apr = calculate_apr_from_monthly(P, A, n)

    if apr == 0.0:
        print("\n⚠️ APR cannot be calculated with the given inputs.")
        print("   Try increasing the monthly payment or loan duration.\n")
    else:
        print(f"\nThe Annual Percentage Rate (APR) is: {apr:.2f}%")
        schedule = generate_amortization_schedule(P, A, monthly_rate, n)
        display_schedule(schedule)
        export_schedule(schedule, P, A, n, apr, output_folder)
